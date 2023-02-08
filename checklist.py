import openpyxl
from openpyxl import load_workbook


checklist = ''
hw2 = '.\code\INFS247HW2.xlsx'
wb = load_workbook(hw2)

chk_file = '.\code\Checks.txt'
with open(chk_file) as f:
    for line in f:
        checklist = checklist + line
checklist = checklist.split('\n')

answers = []

for check in checklist:
    check = check.split(' ')
    sheetname = check[0]
    sheetname = sheetname.replace("_", " ")
    cellrange = check[1]
    arg = check[2]
    value = check[3]
    sheet = wb[sheetname]
    ans = sheet[cellrange]
    try:
        #print(ans.value)
        answers.append(ans.value)
    except:
        for row in sheet.iter_cols(min_row = 3, min_col =1, max_col = 1, max_row = 13, values_only=True):
            answers.append(row)
            #print(row)
    finally:
        print(answers)
            

"""  for row in sheet.iter_cols(min_row = 3, min_col =1, max_col = 1, max_row = 13, values_only=True):
      print(row)
"""


    
